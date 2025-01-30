import openpyxl

def row_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def column_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def read_data(file,sheetname,row_no,column_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row_no,column_no).value

def write_data(file,sheetname,row_no,column_no,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row_no,column_no).value=data
    workbook.save(file)
