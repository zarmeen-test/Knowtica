import openpyxl
path=r'C:\Users\zmuiz\OneDrive\Documents\DDT_data.xlsx'

# #writing the data
# workbook=openpyxl.load_workbook(path)
# Sheet=workbook.active  #workbook['sheetname']
# for r in range(1,4+1):
#     for c in range(1,2+1):
#         Sheet.cell(r,c).value=input('enter the data:')
# workbook.save(path)

#reading the data
workbook=openpyxl.load_workbook(path)
Sheet=workbook.active
rows=Sheet.max_row
columns=Sheet.max_column
print(rows,columns)
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(Sheet.cell(r,c).value,end=' ')
    print()


